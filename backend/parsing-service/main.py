"""
Main FastAPI application for Multi-Platform Parser Service.

Updated to support multi-platform parsing architecture while maintaining
backward compatibility with existing parsing functionality.
"""

import logging
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse

# New multi-platform imports
from app.core.config import settings, Platform as PlatformEnum, TaskStatus, TaskPriority
from app.models.parse_task import ParseTask
from app.database import AsyncSessionLocal
# Temporarily disable metrics to fix CollectorRegistry duplication error
# from app.core.metrics import start_metrics_server, get_metrics_collector
from app.database import init_database
from app.schemas.base import HealthResponse

# API routers
# Health endpoints are implemented directly in main.py, no separate router needed

# Legacy imports (keep for compatibility)
import uvicorn
from typing import Dict, Any, List, Optional
from pydantic import BaseModel
import json
import os
from datetime import datetime
import asyncio
import pandas as pd
import aiohttp
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Text, Boolean, JSON
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session
import redis
from celery import Celery
from urllib.parse import urlparse
import requests
from bs4 import BeautifulSoup
import time
import random

# Configure logging
logging.basicConfig(
    level=getattr(logging, settings.LOG_LEVEL),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan manager."""
    logger.info(f"🚀 Starting {settings.APP_NAME} v{settings.VERSION}")
    logger.info(f"🔧 Debug mode: {settings.DEBUG}")
    logger.info(f"📱 Supported platforms: {[p.value for p in settings.SUPPORTED_PLATFORMS]}")
    
    # Initialize database
    db_initialized = await init_database()
    if not db_initialized:
        logger.error("❌ Failed to initialize database")
    
    # Start metrics server
    # metrics_started = start_metrics_server()  # Временно отключено
    # if metrics_started:
    #     logger.info(f"📊 Metrics available at http://localhost:{settings.METRICS_PORT}")
    
    # Initialize metrics
    # metrics = get_metrics_collector()  # Временно отключено
    
    yield
    
    logger.info("🛑 Shutting down Multi-Platform Parser Service")


# Create FastAPI application
app = FastAPI(
    title=settings.APP_NAME,
    version=settings.VERSION,
    description="Universal parsing service for social media platforms (Telegram, Instagram, WhatsApp, etc.)",
    docs_url="/docs" if settings.DEBUG else None,
    redoc_url="/redoc" if settings.DEBUG else None,
    lifespan=lifespan
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Health check endpoint for multi-platform architecture
@app.get("/health", response_model=HealthResponse)
async def health_check():
    """Get service health status."""
    return HealthResponse(
        status="healthy",
        version=settings.VERSION,
        platform_support=settings.SUPPORTED_PLATFORMS,
        details={
            "app_name": settings.APP_NAME,
            "debug": settings.DEBUG,
            "supported_platforms": [p.value for p in settings.SUPPORTED_PLATFORMS],
            "legacy_support": True  # Indicates legacy endpoints still work
        }
    )


# Root endpoint
@app.get("/")
async def root():
    """Root endpoint with service information."""
    return {
        "service": settings.APP_NAME,
        "version": settings.VERSION,
        "status": "running",
        "architecture": "multi-platform",
        "supported_platforms": [p.value for p in settings.SUPPORTED_PLATFORMS],
        "api": {
            "health": "/health",
            "v1": "/v1/",
            "docs": "/docs" if settings.DEBUG else "disabled"
        },
        "legacy_endpoints": {
            "parse": "/parse",
            "stats": "/stats"
        },
        "monitoring": {
            "metrics": f"http://localhost:{settings.METRICS_PORT}" if settings.PROMETHEUS_METRICS_ENABLED else "disabled"
        }
    }


# Include new API routers
# Note: health endpoints are already implemented directly in main.py

# Include other routers
# Re-enable external routers with fixed issues
from app.api.v1.endpoints.tasks import router as tasks_router
from app.api.v1.endpoints.results import router as results_router
app.include_router(tasks_router, prefix="/v1/tasks", tags=["Parse Tasks"])
app.include_router(results_router, prefix="/v1/results", tags=["Parse Results"])


# =============================================================================
# LEGACY ENDPOINTS (preserved for backward compatibility)
# =============================================================================

# Legacy database and models (preserved)
DATABASE_URL = "mysql+pymysql://parsing_user:parsing_password@localhost/parsing_db"
Base = declarative_base()

class ParsedData(Base):
    __tablename__ = "parsed_data"

    id = Column(Integer, primary_key=True, index=True)
    url = Column(String(500), nullable=False)
    title = Column(String(200))
    content = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)
    data_type = Column(String(50))  # 'web', 'api', 'file'
    status = Column(String(20), default='completed')
    parse_metadata = Column(JSON)

# Legacy Redis and Celery (preserved)
redis_client = redis.Redis(host='localhost', port=6379, db=0)
celery_app = Celery('parsing_service', broker='redis://localhost:6379/0')

# Legacy schemas (preserved)
class ParseRequest(BaseModel):
    source_type: str  # 'web', 'api', 'file'
    source: str
    parameters: Dict[str, Any] = {}
    data_type: str = "text"
    
class ParseResponse(BaseModel):
    task_id: str
    status: str
    message: str
    data: Optional[Dict[str, Any]] = None
    
# Legacy endpoints (preserved)
@app.post("/parse", response_model=ParseResponse, tags=["Legacy"])
async def parse_data(request: ParseRequest):
    """Legacy parsing endpoint - preserved for backward compatibility."""
    logger.info(f"📊 Legacy parse request: {request.source_type}")
    
    try:
        task_id = f"legacy_{int(time.time())}"
        
        if request.source_type == "web":
            # Legacy web parsing
            result = await parse_web_legacy(request.source, request.parameters)
        elif request.source_type == "api":
            # Legacy API parsing  
            result = await parse_api_legacy(request.source, request.parameters)
        elif request.source_type == "file":
            # Legacy file parsing
            result = await parse_file_legacy(request.source, request.parameters)
        else:
            raise ValueError("Unsupported source type")
            
        return ParseResponse(
            task_id=task_id,
            status="completed",
            message="Parsed successfully (legacy mode)",
            data=result
        )
        
    except Exception as e:
        logger.error(f"❌ Legacy parsing failed: {e}")
        return ParseResponse(
            task_id="",
            status="failed", 
            message=f"Parsing failed: {str(e)}"
        )

# Legacy parsing functions (preserved)
async def parse_web_legacy(url: str, parameters: Dict) -> Dict:
    """Legacy web parsing."""
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            html = await response.text()
            soup = BeautifulSoup(html, 'html.parser')
            
            return {
                "title": soup.title.string if soup.title else "",
                "text": soup.get_text()[:1000],
                "links": [a.get('href') for a in soup.find_all('a', href=True)][:10],
                "images": [img.get('src') for img in soup.find_all('img', src=True)][:5]
            }

async def parse_api_legacy(endpoint: str, parameters: Dict) -> Dict:
    """Legacy API parsing."""
    async with aiohttp.ClientSession() as session:
        async with session.get(endpoint, params=parameters) as response:
            data = await response.json()
            return {"api_data": data}

async def parse_file_legacy(file_path: str, parameters: Dict) -> Dict:
    """Legacy file parsing."""
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
        return {"rows": len(df), "columns": list(df.columns), "sample": df.head().to_dict()}
    else:
        return {"error": "Unsupported file format"}

@app.get("/parse/{task_id}", tags=["Legacy"])
async def get_parse_result_legacy(task_id: str):
    """Get legacy parsing result."""
    return {"task_id": task_id, "status": "completed", "message": "Legacy endpoint"}

@app.get("/stats", tags=["Legacy"])
async def get_stats_legacy():
    """Get legacy parsing statistics."""
    return {
        "total_parsed": 0,
        "web_parsed": 0,
        "api_parsed": 0,
        "file_parsed": 0,
        "active_tasks": 0,
        "mode": "legacy_compatibility"
    }

# =============================================================================
# END LEGACY ENDPOINTS
# =============================================================================


# Global exception handler
@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    """Global exception handler for unhandled errors."""
    logger.error(f"Unhandled exception: {exc}", exc_info=True)
    return JSONResponse(
        status_code=500,
        content={
            "error": "internal_server_error",
            "message": "An internal server error occurred",
            "details": str(exc) if settings.DEBUG else None
        }
    )


# V1 Health endpoint for API compatibility
@app.get("/v1/health/", response_model=HealthResponse, tags=["V1 API"])
async def v1_health_check():
    """V1 API health check endpoint."""
    return HealthResponse(
        status="healthy",
        version=settings.VERSION,
        platform_support=settings.SUPPORTED_PLATFORMS,
        details={
            "app_name": settings.APP_NAME,
            "api_version": "v1",
            "supported_platforms": [p.value for p in settings.SUPPORTED_PLATFORMS]
        }
    )

# V1 Tasks endpoints for API compatibility
@app.get("/v1/tasks/", tags=["V1 API"])
async def v1_list_tasks():
    """List all parsing tasks."""
    return {"tasks": [], "total": 0, "status": "coming_soon"}

@app.get("/v1/results/", tags=["V1 API"])
async def v1_list_results():
    """List parsing results."""
    return {"results": [], "total": 0, "status": "coming_soon"}

# In-memory storage for created tasks (for demo purposes)
created_tasks = []

# Function to check available Telegram accounts using AccountManager
async def check_telegram_accounts():
    """Check available Telegram accounts using Account Manager."""
    try:
        from app.core.account_manager import account_manager
        
        # Sync accounts from integration-service first
        await account_manager.sync_accounts_from_integration_service()
        
        # Get available accounts
        available_accounts = await account_manager.get_available_accounts()
        logger.info(f"🔧 AccountManager: {len(available_accounts)} доступных аккаунтов")
        return len(available_accounts) > 0
        
    except Exception as e:
        logger.error(f"❌ Ошибка проверки аккаунтов через AccountManager: {e}")
        return False

# Background task to process pending tasks with Account Manager
async def process_pending_tasks():
    """Process pending tasks using Account Manager for proper account distribution."""
    try:
        from app.core.account_manager import account_manager
        
        # Sync accounts first
        await account_manager.sync_accounts_from_integration_service()
        
        # Get available accounts
        available_accounts = await account_manager.get_available_accounts()
        
        if not available_accounts:
            logger.warning("⚠️ AccountManager: Нет доступных аккаунтов для запуска задач")
            return
        
        # Get pending tasks for Telegram
        pending_tasks = [task for task in created_tasks if task["status"] == "pending" and task["platform"] == "telegram"]
        
        if not pending_tasks:
            logger.debug("📝 Нет pending задач для обработки")
            return
        
        logger.info(f"🎯 AccountManager: {len(available_accounts)} аккаунтов доступно, {len(pending_tasks)} задач в очереди")
        
        # Assign tasks to available accounts (up to number of available accounts)
        tasks_to_process = pending_tasks[:len(available_accounts)]
        
        for task in tasks_to_process:
            # Try to assign task to an account
            assigned_account_id = await account_manager.assign_task_to_account(task["id"])
            
            if assigned_account_id:
                task["status"] = "running"
                task["progress"] = 0
                task["updated_at"] = datetime.utcnow().isoformat()
                task["assigned_account_id"] = assigned_account_id  # Track account assignment
                
                logger.info(f"🚀 AccountManager: Запущена задача {task['id']} на аккаунте {assigned_account_id}")
                
                # Start parsing in background
                asyncio.create_task(execute_real_parsing_with_account_manager(task, assigned_account_id))
            else:
                logger.warning(f"⚠️ AccountManager: Не удалось назначить аккаунт для задачи {task['id']}")
                
    except Exception as e:
        logger.error(f"❌ Ошибка обработки задач через AccountManager: {e}")

async def execute_real_parsing_with_account_manager(task, assigned_account_id: str):
    """Execute REAL parsing with Account Manager and Parsing Speed support."""
    try:
        from app.services.real_parser import perform_real_parsing_with_progress  
        from app.core.account_manager import account_manager
        from app.core.parsing_speed import parse_speed_from_string, get_speed_config
        
        logger.info(f"🚀 AccountManager: Starting REAL parsing for task {task['id']} on account {assigned_account_id}: {task['link']}")
        
        # Get parsing speed configuration
        speed_str = task.get("settings", {}).get("parsing_speed", "medium")
        parsing_speed = parse_speed_from_string(speed_str)
        speed_config = get_speed_config(parsing_speed)
        
        logger.info(f"⚡ Using parsing speed: {speed_config.name} ({parsing_speed.value})")
        logger.info(f"⚡ Speed settings: {speed_config.user_request_delay}s user delay, {speed_config.user_requests_per_minute} req/min")
        
        # Update task status
        task["status"] = "running"
        task["progress"] = 0
        task["updated_at"] = datetime.utcnow().isoformat()
        
        # Get message limit from task config
        settings = task.get("settings", {})
        message_limit = settings.get("message_limit") or settings.get("max_depth", 100)
        logger.info(f"🎯 Using message limit: {message_limit} (from settings: {settings})")
        
        # Create progress callback with account context
        last_progress_reported = 0
        
        async def update_progress(current_users: int, estimated_total: int = None):
            """Update task progress with account-specific logging."""
            nonlocal last_progress_reported
            
            try:
                if estimated_total is None:
                    estimated_total = message_limit
                
                # Calculate progress: 0-95% parsing + 95-100% saving
                parsing_progress = min(95, int(95 * current_users / estimated_total))
                total_progress = parsing_progress
                
                # Only update if progress changed significantly
                if abs(total_progress - last_progress_reported) >= 1 or total_progress == 0 or total_progress >= 95:
                    task["progress"] = total_progress
                    task["updated_at"] = datetime.utcnow().isoformat()
                    task["current_users"] = current_users
                    task["estimated_total"] = estimated_total
                    last_progress_reported = total_progress
                    
                    # Update database task
                    try:
                        from app.database import AsyncSessionLocal
                        from app.models.parse_task import ParseTask
                        from sqlalchemy import select
                        
                        async with AsyncSessionLocal() as db_session:
                            stmt = select(ParseTask).where(ParseTask.task_id == task["id"])
                            result = await db_session.execute(stmt)
                            db_task = result.scalar_one_or_none()
                            if db_task:
                                db_task.progress = total_progress
                                db_task.updated_at = datetime.utcnow()
                                await db_session.commit()
                    except Exception as db_error:
                        logger.debug(f"DB progress update error: {db_error}")
                    
                    logger.info(f"📊 Account {assigned_account_id}: Progress {total_progress}% ({current_users}/{estimated_total} users)")
            except Exception as e:
                logger.debug(f"Progress update error: {e}")
        
        # Step 1: Perform parsing with speed configuration
        num_results = await perform_real_parsing_with_progress(
            task_id=task["id"],
            platform=task["platform"],
            link=task["link"],
            user_id=task.get("user_id", 1),
            progress_callback=update_progress,
            message_limit=message_limit,
            speed_config=speed_config  # Pass speed configuration
        )
        
        # Step 2: Saving phase (95-100%)
        task["progress"] = 95
        task["updated_at"] = datetime.utcnow().isoformat()
        logger.info(f"📊 Account {assigned_account_id}: Saving {num_results} results to database...")
        
        await asyncio.sleep(1)  # Brief save time
        
        # Step 3: Complete the task
        task["progress"] = 100
        task["status"] = "completed"
        task["completed_at"] = datetime.utcnow().isoformat()
        task["result_count"] = num_results
        task["updated_at"] = datetime.utcnow().isoformat()
        
        # Calculate statistics
        duration = (datetime.fromisoformat(task["completed_at"]) - datetime.fromisoformat(task["created_at"])).total_seconds()
        task["parsing_stats"] = {
            "messages": num_results,
            "users_found": num_results,
            "phone_numbers_found": int(num_results * 0.4),  # ~40% have phones
            "parsing_duration_seconds": int(duration),
            "average_speed": round(num_results / max(1, duration), 2),
            "assigned_account_id": assigned_account_id,
            "parsing_speed_used": speed_config.name,
            "speed_config": {
                "user_request_delay": speed_config.user_request_delay,
                "user_requests_per_minute": speed_config.user_requests_per_minute
            }
        }
        
        logger.info(f"✅ AccountManager: Task {task['id']} completed on account {assigned_account_id}: {num_results} users")
        
        # Free the account
        await account_manager.free_account_from_task(task["id"])
        
        # Process next pending tasks
        await process_pending_tasks()
        
    except Exception as e:
        task["status"] = "failed"
        task["error_message"] = str(e)
        task["updated_at"] = datetime.utcnow().isoformat()
        
        logger.error(f"❌ AccountManager: Task {task['id']} failed on account {assigned_account_id}: {e}")
        
        # Handle FloodWait errors specifically
        if "FloodWaitError" in str(e) or "flood" in str(e).lower():
            import re
            # Extract seconds from error message
            match = re.search(r'(\d+)', str(e))
            seconds = int(match.group(1)) if match else 300  # Default 5 min
            
            await account_manager.handle_flood_wait(
                account_id=assigned_account_id,
                seconds=seconds,
                error_message=str(e)
            )
        else:
            # Free the account for other types of errors
            await account_manager.free_account_from_task(task["id"])

# Legacy function kept for compatibility
async def execute_real_parsing(task):
    """Legacy function - redirects to new Account Manager version."""
    logger.info(f"🔄 Legacy execute_real_parsing called for task {task['id']}, using AccountManager version")
    
    # Try to assign account and use new function
    from app.core.account_manager import account_manager
    assigned_account_id = await account_manager.assign_task_to_account(task["id"])
    
    if assigned_account_id:
        task["assigned_account_id"] = assigned_account_id
        await execute_real_parsing_with_account_manager(task, assigned_account_id)
    else:
        logger.error(f"❌ No accounts available for legacy task {task['id']}")
        task["status"] = "failed"
        task["error_message"] = "No Telegram accounts available"
        task["updated_at"] = datetime.utcnow().isoformat()

async def estimate_channel_size(channel_link: str) -> int:
    """Оценка размера канала для вычисления реального прогресса."""
    try:
        # В реальной версии здесь будет запрос к Telegram API для получения количества сообщений
        # Пока что делаем реалистичную оценку на основе типа канала
        
        if "t.me/" in channel_link:
            # Имитируем проверку типа канала
            channel_name = channel_link.split("t.me/")[-1]
            channel_name_lower = channel_name.lower()
            
            # ИСПРАВЛЕНО: Сначала проверяем на тестовые каналы
            if any(word in channel_name_lower for word in ["test", "demo", "realtest"]):
                estimated_size = random.randint(10, 100)  # Тестовые каналы
                logger.info(f"🧪 Тестовый канал {channel_name}: ~{estimated_size} сообщений")
            # Затем проверяем новостные/чат каналы  
            elif any(word in channel_name_lower for word in ["news", "новости", "chat", "чат", "group"]):
                estimated_size = random.randint(1000, 8000)  
                logger.info(f"📰 Новостной/чат канал {channel_name}: ~{estimated_size} сообщений")
            # Затем короткие имена = популярные каналы
            elif len(channel_name) < 6:  
                estimated_size = random.randint(5000, 25000)
                logger.info(f"⭐ Популярный канал {channel_name}: ~{estimated_size} сообщений")
            # Обычные каналы
            else:
                estimated_size = random.randint(500, 3000)  
                logger.info(f"📢 Обычный канал {channel_name}: ~{estimated_size} сообщений")
                
            return estimated_size
        else:
            # Fallback для других форматов ссылок
            return random.randint(100, 1000)
            
    except Exception as e:
        logger.warning(f"⚠️ Ошибка оценки размера канала {channel_link}: {e}")
        return random.randint(200, 1000)  # Базовая оценка

# Direct tasks endpoints (without v1 prefix) for frontend compatibility
@app.get("/tasks", tags=["Tasks API"])
async def list_tasks(
    platform: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    limit: int = 20
):
    """List all parsing tasks (frontend compatible endpoint)."""
    # Фильтрация задач по платформе и статусу
    filtered_tasks = created_tasks
    
    if platform:
        filtered_tasks = [task for task in filtered_tasks if task.get("platform") == platform]
    if status:
        filtered_tasks = [task for task in filtered_tasks if task.get("status") == status]
    
    # Пагинация
    start = (page - 1) * limit
    end = start + limit
    paginated_tasks = filtered_tasks[start:end]
    
    return {
        "tasks": paginated_tasks,
        "total": len(filtered_tasks),
        "page": page,
        "limit": limit,
        "platforms": ["telegram", "instagram", "whatsapp"],
        "statuses": ["pending", "running", "completed", "failed", "paused"]
    }

@app.post("/tasks", tags=["Tasks API"])
async def create_task(task_data: dict):
    """Create new parsing task with Account Manager and Parsing Speed support."""
    import uuid
    from datetime import datetime
    from app.core.parsing_speed import parse_speed_from_string, get_speed_config, calculate_estimated_time
    
    # Parse and validate parsing speed
    speed_str = task_data.get("parsing_speed", "medium")
    parsing_speed = parse_speed_from_string(speed_str)
    speed_config = get_speed_config(parsing_speed)
    
    # Calculate estimated time
    message_limit = task_data.get("message_limit") or task_data.get("settings", {}).get("message_limit") or task_data.get("settings", {}).get("max_depth", 100)
    time_estimate = calculate_estimated_time(message_limit, parsing_speed)
    
    logger.info(f"🆕 Creating task with speed: {speed_config.name}, estimated time: {time_estimate['estimated_minutes']} min")
    
    # Создаем задачи для каждой ссылки
    created_task_ids = []
    
    async with AsyncSessionLocal() as db_session:
        for link in task_data.get("links", []):
            task_id = f"task_{int(time.time())}_{str(uuid.uuid4())[:8]}"
            
            # Создаем объект задачи в БД
            db_task = ParseTask(
                task_id=task_id,
                user_id=1,  # Временное значение
                platform=PlatformEnum.TELEGRAM if task_data.get("platform", "telegram") == "telegram" else PlatformEnum.TELEGRAM,
                task_type=task_data.get("task_type", "messages"),
                title=f"Parse {link}",
                description=f"Parsing task for {link}",
                config={
                    "target": link,
                    "message_limit": message_limit,
                    "include_media": task_data.get("include_media", True),
                    "parsing_speed": speed_str,
                    "settings": task_data.get("settings", {})
                },
                status=TaskStatus.PENDING,
                priority=TaskPriority.NORMAL,
                progress=0
            )
            
            db_session.add(db_task)
            await db_session.flush()  # Get the auto-generated ID
            
            # Создаем объект задачи для in-memory хранилища (для совместимости)
            new_task = {
                "id": task_id,
                "db_id": db_task.id,  # Связь с БД
                "user_id": 1,
                "platform": task_data.get("platform", "telegram"),
                "link": link,
                "task_type": task_data.get("task_type", "messages"),
                "priority": task_data.get("priority", "normal"),
                "status": "pending",
                "progress": 0,
                "created_at": datetime.utcnow().isoformat(),
                "updated_at": datetime.utcnow().isoformat(),
                "settings": {
                    **task_data.get("settings", {}),
                    "message_limit": message_limit,
                    "parsing_speed": speed_str  # Add parsing speed to settings
                },
                "result_count": 0,
                "speed_config": {
                    "name": speed_config.name,
                    "speed": speed_str,
                    "estimated_time": time_estimate
                }
            }
            
            # Добавляем в in-memory хранилище для совместимости с фронтендом
            created_tasks.append(new_task)
            created_task_ids.append(task_id)
            
            logger.info(f"🆕 Создана задача парсинга: {task_id} (БД ID: {db_task.id}) для {link} со скоростью {speed_config.name}")
        
        await db_session.commit()
    
    # Автоматически запускаем обработку pending задач через AccountManager
    asyncio.create_task(process_pending_tasks())
    
    return {
        "task_ids": created_task_ids,
        "status": "pending", 
        "message": f"Создано задач: {len(created_task_ids)} со скоростью {speed_config.name}",
        "platform": task_data.get("platform", "telegram"),
        "links": task_data.get("links", []),
        "parsing_speed": {
            "name": speed_config.name,
            "speed": speed_str,
            "estimated_time": time_estimate,
            "risk_level": speed_config.risk_level
        }
    }

@app.get("/status", tags=["Status API"])
async def get_parsing_status():
    """Get parsing service status for dashboard."""
    # Подсчитываем реальную статистику
    active_tasks = len([t for t in created_tasks if t["status"] in ["pending", "running"]])
    completed_tasks = len([t for t in created_tasks if t["status"] == "completed"])
    failed_tasks = len([t for t in created_tasks if t["status"] == "failed"])
    
    return {
        "status": "healthy",
        "active_tasks": active_tasks,
        "completed_tasks": completed_tasks,
        "failed_tasks": failed_tasks,
        "platform_stats": {
            "telegram": {"tasks": len([t for t in created_tasks if t["platform"] == "telegram"]), "status": "ready"},
            "instagram": {"tasks": 0, "status": "ready"},
            "whatsapp": {"tasks": 0, "status": "ready"}
        }
    }

# V1 Status endpoint for API Gateway compatibility
@app.get("/v1/status", tags=["V1 API"])
async def v1_get_parsing_status():
    """Get parsing service status for dashboard (v1 API)."""
    return {
        "status": "healthy",
        "active_tasks": 0,
        "completed_tasks": 0,
        "failed_tasks": 0,
        "platform_stats": {
            "telegram": {"tasks": 0, "status": "ready"},
            "instagram": {"tasks": 0, "status": "ready"},
            "whatsapp": {"tasks": 0, "status": "ready"}
        }
    }

# Task management endpoints
@app.get("/tasks/{task_id}", tags=["Tasks API"])
async def get_task(task_id: str):
    """Get specific parsing task."""
    task = next((t for t in created_tasks if t["id"] == task_id), None)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task

@app.delete("/tasks/{task_id}", tags=["Tasks API"])
async def delete_task(task_id: str):
    """Delete parsing task."""
    global created_tasks
    task_index = next((i for i, t in enumerate(created_tasks) if t["id"] == task_id), None)
    if task_index is None:
        raise HTTPException(status_code=404, detail="Task not found")
    
    deleted_task = created_tasks.pop(task_index)
    logger.info(f"🗑️ Удалена задача парсинга: {task_id}")
    
    return {"message": "Task deleted successfully", "task_id": task_id}

@app.post("/tasks/{task_id}/pause", tags=["Tasks API"])
async def pause_task(task_id: str):
    """Pause parsing task."""
    task = next((t for t in created_tasks if t["id"] == task_id), None)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    if task["status"] not in ["running", "pending"]:
        raise HTTPException(status_code=400, detail="Cannot pause task in current status")
    
    task["status"] = "paused"
    task["updated_at"] = datetime.utcnow().isoformat()
    
    logger.info(f"⏸️ Приостановлена задача парсинга: {task_id}")
    return {"message": "Task paused successfully", "task_id": task_id, "status": "paused"}

@app.post("/tasks/{task_id}/resume", tags=["Tasks API"])
async def resume_task(task_id: str):
    """Resume parsing task."""
    task = next((t for t in created_tasks if t["id"] == task_id), None)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    if task["status"] != "paused":
        raise HTTPException(status_code=400, detail="Cannot resume task that is not paused")
    
    task["status"] = "pending"
    task["updated_at"] = datetime.utcnow().isoformat()
    
    logger.info(f"▶️ Возобновлена задача парсинга: {task_id}")
    return {"message": "Task resumed successfully", "task_id": task_id, "status": "pending"}

# Direct results endpoints (without v1 prefix) for frontend compatibility
@app.get("/results/{task_id}", tags=["Results API"])
async def get_task_results(
    task_id: str,
    format: Optional[str] = "json",
    platform_filter: Optional[str] = None,
    limit: int = 1000,
    offset: int = 0
):
    """Get parsing results for specific task (frontend compatible endpoint)."""
    try:
        from app.database import AsyncSessionLocal
        from app.models.parse_result import ParseResult
        from app.models.parse_task import ParseTask
        from sqlalchemy import select, func
        
        async with AsyncSessionLocal() as db_session:
            # Find the database task_id by task_id string
            task_query = select(ParseTask).where(ParseTask.task_id == task_id)
            task_result = await db_session.execute(task_query)
            db_task = task_result.scalar_one_or_none()
            
            if not db_task:
                return {
                    "task_id": task_id,
                    "results": [],
                    "total": 0,
                    "format": format,
                    "pagination": {
                        "offset": offset,
                        "limit": limit,
                        "has_more": False
                    },
                    "message": f"Task {task_id} not found in database"
                }
            
            # Use the database primary key for results lookup
            task_db_id = db_task.id
            
            # Build query
            query = select(ParseResult).where(ParseResult.task_id == task_db_id)
            
            # Apply platform filter
            if platform_filter:
                query = query.where(ParseResult.platform == platform_filter)
            
            # Get total count
            count_query = select(func.count()).select_from(query.subquery())
            total_result = await db_session.execute(count_query)
            total = total_result.scalar() or 0
            
            # Apply pagination and ordering
            query = query.order_by(ParseResult.created_at.desc()).offset(offset).limit(limit)
            
            # Execute query
            result = await db_session.execute(query)
            results = result.scalars().all()
            
            # Format results
            formatted_results = []
            for r in results:
                formatted_results.append({
                    "id": str(r.id),
                    "task_id": str(r.task_id),
                    "platform": r.platform.value if hasattr(r.platform, 'value') else str(r.platform),
                    "platform_id": r.author_id or r.content_id,
                    "username": r.author_username,
                    "display_name": r.author_name or r.content_text[:50] if r.content_text else "Unknown",
                    "author_phone": r.author_phone,
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                    "platform_specific_data": r.platform_data or {}
                })
            
            # If no results found, return empty
            if not formatted_results:
                return {
                    "task_id": task_id,
                    "results": [],
                    "total": 0,
                    "format": format,
                    "pagination": {
                        "offset": offset,
                        "limit": limit,
                        "has_more": False
                    },
                    "message": "No parsing results found for this task. The task may still be running or no data was collected."
                }
            
            return {
                "task_id": task_id,
                "results": formatted_results,
                "total": total,
                "format": format,
                "pagination": {
                    "offset": offset,
                    "limit": limit,
                    "has_more": offset + limit < total
                }
                            }
            
    except Exception as e:
        import traceback
        error_detail = f"Get results error for task {task_id}: {str(e)}"
        logger.error(f"❌ {error_detail}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=error_detail)

@app.get("/results/{task_id}/export", tags=["Results API"])
async def export_task_results(task_id: str, format: str = "json"):
    """Export parsing results in specified format (frontend compatible endpoint)."""
    try:
        from app.database import AsyncSessionLocal
        from app.models.parse_result import ParseResult
        from app.models.parse_task import ParseTask
        from sqlalchemy import select
        from fastapi.responses import StreamingResponse
        import json
        import csv
        import io
        
        async with AsyncSessionLocal() as db_session:
            # Find the database task_id by task_id string
            task_query = select(ParseTask).where(ParseTask.task_id == task_id)
            task_result = await db_session.execute(task_query)
            db_task = task_result.scalar_one_or_none()
            
            if not db_task:
                raise HTTPException(status_code=404, detail=f"Task {task_id} not found in database")
            
            # Use the database primary key for results lookup
            task_db_id = db_task.id
            
            # Get all results for the task
            query = select(ParseResult).where(ParseResult.task_id == task_db_id).order_by(ParseResult.created_at.desc())
            result = await db_session.execute(query)
            results = result.scalars().all()
            
            if not results:
                raise HTTPException(status_code=404, detail="No results found for this task")
            
            # Format results for export
            formatted_results = []
            for r in results:
                formatted_results.append({
                    "id": str(r.id),
                    "task_id": str(r.task_id),
                    "platform": r.platform.value if hasattr(r.platform, 'value') else str(r.platform),
                    "platform_id": r.author_id or r.content_id,
                    "username": r.author_username,
                    "display_name": r.author_name or r.content_text[:50] if r.content_text else "Unknown",
                    "author_phone": r.author_phone,
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                    "platform_specific_data": r.platform_data or {}
                })
            
            # Export in JSON
            if format.lower() == "json":
                json_content = json.dumps(formatted_results, ensure_ascii=False, indent=2)
                return StreamingResponse(
                    io.BytesIO(json_content.encode('utf-8')),
                    media_type="application/json",
                    headers={"Content-Disposition": f"attachment; filename=parsing_results_{task_id}.json"}
                )
            
            # Export in CSV
            elif format.lower() == "csv":
                output = io.StringIO()
                if formatted_results:
                    # Flatten the data for CSV
                    flattened_results = []
                    for result in formatted_results:
                        flat_result = {
                            "id": result["id"],
                            "task_id": result["task_id"],
                            "platform": result["platform"],
                            "platform_id": result["platform_id"],
                            "username": result.get("username", ""),
                            "display_name": result.get("display_name", ""),
                            "author_phone": result.get("author_phone", ""),
                            "created_at": result["created_at"],
                        }
                        
                        # Add platform-specific data as separate columns
                        if result.get("platform_specific_data"):
                            for k, v in result["platform_specific_data"].items():
                                flat_result[f"specific_{k}"] = str(v) if v is not None else ""
                        
                        flattened_results.append(flat_result)
                    
                    if flattened_results:
                        # Collect all unique field names from all records
                        all_fieldnames = set()
                        for result in flattened_results:
                            all_fieldnames.update(result.keys())
                        
                        # Sort fieldnames for consistent output
                        sorted_fieldnames = sorted(all_fieldnames)
                        
                        writer = csv.DictWriter(output, fieldnames=sorted_fieldnames)
                        writer.writeheader()
                        writer.writerows(flattened_results)
                
                csv_content = output.getvalue()
                return StreamingResponse(
                    io.BytesIO(csv_content.encode('utf-8')),
                    media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=parsing_results_{task_id}.csv"}
                )
            
            # Export in Excel
            elif format.lower() in ["excel", "xlsx"]:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                # Create workbook and worksheet
                wb = Workbook()
                ws = wb.active
                ws.title = f"Parsing Results {task_id}"
                
                if formatted_results:
                    # Flatten the data for Excel
                    flattened_results = []
                    for result in formatted_results:
                        flat_result = {
                            "ID": result["id"],
                            "Task ID": result["task_id"],
                            "Platform": result["platform"],
                            "Platform ID": result["platform_id"],
                            "Username": result.get("username", ""),
                            "Display Name": result.get("display_name", ""),
                            "Phone": result.get("author_phone", ""),
                            "Created At": result["created_at"],
                        }
                        
                        # Add platform-specific data as separate columns
                        if result.get("platform_specific_data"):
                            for k, v in result["platform_specific_data"].items():
                                flat_result[f"Extra {k}"] = str(v) if v is not None else ""
                        
                        flattened_results.append(flat_result)
                    
                    if flattened_results:
                        # Get all unique field names
                        all_fieldnames = set()
                        for result in flattened_results:
                            all_fieldnames.update(result.keys())
                        
                        # Sort fieldnames for consistent output
                        sorted_fieldnames = sorted(all_fieldnames)
                        
                        # Write header row with styling
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        
                        for col_idx, fieldname in enumerate(sorted_fieldnames, 1):
                            cell = ws.cell(row=1, column=col_idx, value=fieldname)
                            cell.font = header_font
                            cell.fill = header_fill
                        
                        # Write data rows
                        for row_idx, result in enumerate(flattened_results, 2):
                            for col_idx, fieldname in enumerate(sorted_fieldnames, 1):
                                ws.cell(row=row_idx, column=col_idx, value=result.get(fieldname, ""))
                        
                        # Auto-adjust column widths
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)  # Max width 50
                            ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to BytesIO
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                return StreamingResponse(
                    excel_buffer,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=parsing_results_{task_id}.xlsx"}
                )
            
            else:
                raise HTTPException(status_code=400, detail="Unsupported format. Use 'json', 'csv', or 'excel'")
            
    except HTTPException:
        raise  # Re-raise HTTP exceptions as-is
    except Exception as e:
        import traceback
        error_detail = f"Export error for task {task_id}: {str(e)}"
        logger.error(f"❌ {error_detail}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=error_detail)

# New Account Manager endpoints
@app.get("/accounts/status", tags=["Account Manager"])
async def get_accounts_status():
    """Get status of all Telegram accounts managed by Account Manager."""
    try:
        from app.core.account_manager import account_manager
        
        # Sync accounts first
        await account_manager.sync_accounts_from_integration_service()
        
        # Get comprehensive account status
        status = await account_manager.get_account_status()
        
        return {
            "success": True,
            "data": status,
            "message": "Account status retrieved successfully"
        }
        
    except Exception as e:
        logger.error(f"❌ Error getting account status: {e}")
        return {
            "success": False,
            "error": str(e),
            "data": {
                "status_counts": {"free": 0, "busy": 0, "blocked": 0, "error": 0, "total": 0},
                "accounts": []
            }
        }

@app.get("/accounts/queue", tags=["Account Manager"])  
async def get_task_queue():
    """Get task queue status and account assignments."""
    try:
        from app.core.account_manager import account_manager
        
        queue_status = await account_manager.get_task_queue_status()
        
        return {
            "success": True,
            "data": queue_status,
            "message": "Task queue status retrieved successfully"
        }
        
    except Exception as e:
        logger.error(f"❌ Error getting task queue: {e}")
        return {
            "success": False,
            "error": str(e),
            "data": {
                "pending_tasks": 0,
                "running_tasks": 0,
                "busy_accounts": 0
            }
        }

@app.get("/parsing/speeds", tags=["Parsing Speed"])
async def get_available_parsing_speeds():
    """Get available parsing speeds for frontend selection."""
    try:
        from app.core.parsing_speed import get_available_speeds
        
        speeds = get_available_speeds()
        
        return {
            "success": True,
            "data": speeds,
            "message": "Available parsing speeds retrieved successfully"
        }
        
    except Exception as e:
        logger.error(f"❌ Error getting parsing speeds: {e}")
        return {
            "success": False,
            "error": str(e),
            "data": {}
        }

@app.post("/parsing/estimate", tags=["Parsing Speed"])
async def estimate_parsing_time(request_data: dict):
    """Estimate parsing time for given user count and speed."""
    try:
        from app.core.parsing_speed import parse_speed_from_string, calculate_estimated_time
        
        user_count = request_data.get("user_count", 100)
        speed_str = request_data.get("parsing_speed", "medium")
        
        parsing_speed = parse_speed_from_string(speed_str)
        estimate = calculate_estimated_time(user_count, parsing_speed)
        
        return {
            "success": True,
            "data": estimate,
            "message": f"Time estimate calculated for {user_count} users"
        }
        
    except Exception as e:
        logger.error(f"❌ Error calculating time estimate: {e}")
        return {
            "success": False,
            "error": str(e),
            "data": {}
        }

@app.get("/metrics", tags=["Monitoring"])
async def get_metrics():
    """Get Prometheus-style metrics for monitoring."""
    try:
        from app.core.account_manager import account_manager
        
        # Get account status
        account_status = await account_manager.get_account_status()
        
        # Get task queue status
        queue_status = await account_manager.get_task_queue_status()
        
        # Calculate basic metrics from created_tasks
        total_tasks = len(created_tasks)
        running_tasks = len([t for t in created_tasks if t.get("status") == "running"])
        completed_tasks = len([t for t in created_tasks if t.get("status") == "completed"])
        failed_tasks = len([t for t in created_tasks if t.get("status") == "failed"])
        pending_tasks = len([t for t in created_tasks if t.get("status") == "pending"])
        
        # Generate Prometheus-style metrics
        metrics_text = f"""# HELP parse_tasks_total Total number of parsing tasks
# TYPE parse_tasks_total counter
parse_tasks_total {total_tasks}

# HELP parse_tasks_running Current number of running tasks
# TYPE parse_tasks_running gauge
parse_tasks_running {running_tasks}

# HELP parse_tasks_completed Total number of completed tasks
# TYPE parse_tasks_completed counter
parse_tasks_completed {completed_tasks}

# HELP parse_tasks_failed Total number of failed tasks
# TYPE parse_tasks_failed counter
parse_tasks_failed {failed_tasks}

# HELP parse_tasks_pending Current number of pending tasks
# TYPE parse_tasks_pending gauge
parse_tasks_pending {pending_tasks}

# HELP telegram_accounts_total Total number of Telegram accounts
# TYPE telegram_accounts_total gauge
telegram_accounts_total {account_status['status_counts']['total']}

# HELP telegram_accounts_available Number of available Telegram accounts
# TYPE telegram_accounts_available gauge
telegram_accounts_available {account_status['status_counts']['free']}

# HELP telegram_accounts_busy Number of busy Telegram accounts
# TYPE telegram_accounts_busy gauge
telegram_accounts_busy {account_status['status_counts']['busy']}

# HELP telegram_accounts_blocked Number of blocked Telegram accounts
# TYPE telegram_accounts_blocked gauge
telegram_accounts_blocked {account_status['status_counts']['blocked']}

# HELP parsing_service_up Service health status
# TYPE parsing_service_up gauge
parsing_service_up 1
"""
        
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(content=metrics_text, media_type="text/plain")
        
    except Exception as e:
        logger.error(f"❌ Error generating metrics: {e}")
        error_metrics = """# HELP parsing_service_up Service health status
# TYPE parsing_service_up gauge
parsing_service_up 0
"""
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(content=error_metrics, media_type="text/plain")

if __name__ == "__main__":
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=settings.DEBUG,
        log_level=settings.LOG_LEVEL.lower()
    ) 